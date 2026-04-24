from __future__ import annotations

import csv
import io
import json
import logging
import os
import threading
import uuid
from dataclasses import dataclass
from datetime import datetime
from datetime import timezone
from pathlib import Path
from typing import Any
from urllib import error as urllib_error
from urllib import request as urllib_request

import boto3
import redis
from flask import Flask
from flask import jsonify
from flask import request
from flask import send_file
from sqlalchemy import create_engine
from sqlalchemy import select
from sqlalchemy import text
from sqlalchemy import update
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.orm import Session
from sqlalchemy.orm import sessionmaker
from werkzeug.utils import secure_filename

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency
    load_workbook = None

from db.models import EmailCache
from db.models import Job
from db.models import User
from services.validator import EmailValidator
from services.validator import ValidationResult


LOGGER = logging.getLogger(__name__)

ALLOWED_EXTENSIONS = {".csv", ".txt", ".xlsx"}
EMAIL_RESULT_TTL_SECONDS = 60 * 60 * 24 * 30
TRANSIENT_EMAIL_RESULT_TTL_SECONDS = 60 * 15
PROGRESS_TTL_SECONDS = 60 * 60 * 24 * 30
EMAIL_CACHE_PREFIX = "email-cache:"
JOB_PROGRESS_PREFIX = "job-progress:"
JOB_CANCEL_PREFIX = "job-cancel:"
EMAIL_REGEX = EmailValidator.SYNTAX_REGEX
TRANSIENT_RISKY_REASONS = {
    "smtp_timeout",
    "probe_timeout",
    "probe_unreachable",
    "probe_error",
    "domain_rate_limited",
    "daily_limit_reached",
}


@dataclass(slots=True)
class BackendServices:
    session_factory: sessionmaker[Session]
    redis_client: redis.Redis
    r2_client: Any
    r2_bucket: str
    validator: EmailValidator
    probe_health_url: str


def create_app() -> Flask:
    app = Flask(__name__)
    app.config["DEFAULT_USER_ID"] = os.getenv("DEFAULT_USER_ID")
    app.config["UPLOAD_MAX_ROWS"] = int(os.getenv("UPLOAD_MAX_ROWS", "1000000"))
    app.extensions["backend_services"] = build_services()
    register_routes(app)
    return app


def build_services() -> BackendServices:
    database_url = require_env("DATABASE_URL")
    redis_url = require_env("REDIS_URL")
    r2_bucket = require_env("R2_BUCKET")
    probe_server_ip = os.getenv("PROBE_SERVER_IP", "84.8.217.135")
    probe_server_port = int(os.getenv("PROBE_SERVER_PORT", "8080"))

    engine = create_engine(database_url, pool_pre_ping=True)
    session_factory = sessionmaker(bind=engine, expire_on_commit=False)

    redis_client = redis.Redis.from_url(redis_url, decode_responses=True)

    r2_endpoint_url = os.getenv("R2_ENDPOINT_URL")
    if not r2_endpoint_url:
        account_id = require_env("R2_ACCOUNT_ID")
        r2_endpoint_url = f"https://{account_id}.r2.cloudflarestorage.com"

    r2_client = boto3.client(
        "s3",
        endpoint_url=r2_endpoint_url,
        aws_access_key_id=require_env("R2_ACCESS_KEY_ID"),
        aws_secret_access_key=require_env("R2_SECRET_ACCESS_KEY"),
        region_name=os.getenv("R2_REGION", "auto"),
    )

    validator = EmailValidator(
        probe_server_ip=probe_server_ip,
        probe_server_port=probe_server_port,
        redis_client=redis_client,
    )

    return BackendServices(
        session_factory=session_factory,
        redis_client=redis_client,
        r2_client=r2_client,
        r2_bucket=r2_bucket,
        validator=validator,
        probe_health_url=f"http://{probe_server_ip}:{probe_server_port}/health",
    )


def register_routes(app: Flask) -> None:
    @app.post("/api/upload")
    def upload() -> object:
        services = get_services(app)
        uploaded_file = request.files.get("file")
        if uploaded_file is None:
            return jsonify({"error": "file is required"}), 400

        filename = secure_filename(uploaded_file.filename or "")
        extension = Path(filename).suffix.lower()
        if extension not in ALLOWED_EXTENSIONS:
            return jsonify({"error": "unsupported file type"}), 400

        try:
            user_id = resolve_user_id(app, services)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

        file_bytes = uploaded_file.read()
        try:
            parsed_emails = parse_uploaded_emails(
                file_bytes=file_bytes,
                extension=extension,
                max_rows=app.config["UPLOAD_MAX_ROWS"],
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

        deduplicated_emails = services.validator.deduplicate_emails(parsed_emails)
        if not deduplicated_emails:
            return jsonify({"error": "no emails found in file"}), 400

        job_id = uuid.uuid4()
        with services.session_factory() as session:
            job = Job(
                id=job_id,
                user_id=user_id,
                filename=filename,
                total_emails=len(deduplicated_emails),
                processed=0,
                valid_count=0,
                risky_count=0,
                invalid_count=0,
                status="queued",
            )
            session.add(job)
            session.commit()

        store_progress(
            services.redis_client,
            job_id,
            build_progress_payload(
                total=len(deduplicated_emails),
                processed=0,
                valid=0,
                risky=0,
                invalid=0,
                status="queued",
            ),
        )
        services.redis_client.delete(cancel_key(job_id))

        worker = threading.Thread(
            target=process_job,
            args=(app, job_id, deduplicated_emails),
            daemon=True,
            name=f"job-{job_id}",
        )
        worker.start()

        return jsonify({"job_id": str(job_id)})

    @app.get("/api/progress")
    def progress() -> object:
        services = get_services(app)
        job_id = parse_job_id(request.args.get("job_id"))
        if job_id is None:
            return jsonify({"error": "valid job_id is required"}), 400

        cached_progress = load_progress(services.redis_client, job_id)
        if cached_progress is not None:
            return jsonify(progress_response_payload(cached_progress))

        with services.session_factory() as session:
            job = session.get(Job, job_id)
            if job is None:
                return jsonify({"error": "job not found"}), 404

            payload = build_progress_payload(
                total=job.total_emails,
                processed=job.processed,
                valid=job.valid_count,
                risky=job.risky_count,
                invalid=job.invalid_count,
                status=job.status,
            )
            return jsonify(progress_response_payload(payload))

    @app.get("/api/download")
    def download() -> object:
        services = get_services(app)
        job_id = parse_job_id(request.args.get("job_id"))
        if job_id is None:
            return jsonify({"error": "valid job_id is required"}), 400

        download_type = (request.args.get("type") or "full").strip().lower()
        if download_type not in {"safe", "full", "risky", "invalid"}:
            return jsonify({"error": "invalid download type"}), 400

        with services.session_factory() as session:
            job = session.get(Job, job_id)
            if job is None:
                return jsonify({"error": "job not found"}), 404
            if not job.r2_file_key:
                return jsonify({"error": "result file not ready"}), 409
            r2_file_key = job.r2_file_key

        try:
            source_bytes = fetch_r2_object(services, r2_file_key)
        except Exception as exc:  # pragma: no cover - network/storage dependent
            LOGGER.exception("Failed to fetch result file for job %s", job_id)
            return jsonify({"error": f"failed to fetch result file: {exc}"}), 502

        filtered_csv = filter_result_csv(source_bytes, download_type)
        download_name = f"{download_type}-emails-{job_id}.csv"
        return send_file(
            io.BytesIO(filtered_csv),
            mimetype="text/csv",
            as_attachment=True,
            download_name=download_name,
        )

    @app.post("/api/cancel")
    def cancel() -> tuple[str, int]:
        services = get_services(app)
        job_id = parse_job_id(request.args.get("job_id"))
        if job_id is None:
            return "", 400

        services.redis_client.setex(cancel_key(job_id), PROGRESS_TTL_SECONDS, "1")
        return "", 204

    @app.get("/api/health")
    def health() -> object:
        services = get_services(app)
        server_status = {
            "status": "ok",
            "database": "ok",
            "redis": "ok",
            "r2": "ok",
        }

        try:
            with services.session_factory() as session:
                session.execute(text("select 1"))
        except Exception as exc:  # pragma: no cover - environment dependent
            server_status["status"] = "degraded"
            server_status["database"] = f"error: {exc}"

        try:
            services.redis_client.ping()
        except Exception as exc:  # pragma: no cover - environment dependent
            server_status["status"] = "degraded"
            server_status["redis"] = f"error: {exc}"

        try:
            services.r2_client.head_bucket(Bucket=services.r2_bucket)
        except Exception as exc:  # pragma: no cover - environment dependent
            server_status["status"] = "degraded"
            server_status["r2"] = f"error: {exc}"

        probe_status = fetch_probe_health(services.probe_health_url)
        if probe_status.get("status") == "error":
            server_status["status"] = "degraded"

        return jsonify({"server": server_status, "probe_server": probe_status})


def process_job(app: Flask, job_id: uuid.UUID, emails: list[str]) -> None:
    services = get_services(app)
    progress = build_progress_payload(
        total=len(emails),
        processed=0,
        valid=0,
        risky=0,
        invalid=0,
        status="running",
    )
    results: list[dict[str, str]] = []

    with app.app_context():
        try:
            with services.session_factory() as session:
                set_job_status(session, job_id, "running")
                store_progress(services.redis_client, job_id, progress)

                for email in emails:
                    if is_cancel_requested(services.redis_client, job_id):
                        finalize_job(
                            services=services,
                            session=session,
                            job_id=job_id,
                            results=results,
                            progress=progress,
                            final_status="cancelled",
                        )
                        return

                    result = load_cached_email_result(services.redis_client, email)
                    if result is None:
                        result = services.validator.validate_email(email)
                        store_cached_email_result(services.redis_client, result)
                        upsert_email_cache(session, result)

                    results.append(
                        {
                            "email": result.email,
                            "status": result.status,
                            "reason": result.reason,
                        }
                    )
                    update_progress_counts(progress, result.status)
                    store_progress(services.redis_client, job_id, progress)

                    if progress["processed"] % 25 == 0 or progress["processed"] == progress["total"]:
                        sync_job_progress(session, job_id, progress)

                finalize_job(
                    services=services,
                    session=session,
                    job_id=job_id,
                    results=results,
                    progress=progress,
                    final_status="done",
                )
        except Exception:  # pragma: no cover - integration path
            LOGGER.exception("Job %s failed during processing", job_id)
            progress["status"] = "cancelled"
            store_progress(services.redis_client, job_id, progress)
            try:
                with services.session_factory() as session:
                    sync_job_progress(session, job_id, progress, status="cancelled")
            except Exception:
                LOGGER.exception("Failed to persist cancelled state for job %s", job_id)


def finalize_job(
    services: BackendServices,
    session: Session,
    job_id: uuid.UUID,
    results: list[dict[str, str]],
    progress: dict[str, Any],
    final_status: str,
) -> None:
    csv_bytes = render_results_csv(results)
    r2_file_key = f"jobs/{job_id}/results.csv"
    upload_r2_object(services, r2_file_key, csv_bytes)

    progress["status"] = final_status
    store_progress(services.redis_client, job_id, progress)
    sync_job_progress(
        session,
        job_id,
        progress,
        status=final_status,
        r2_file_key=r2_file_key,
        finished_at=datetime.now(timezone.utc),
    )


def parse_uploaded_emails(file_bytes: bytes, extension: str, max_rows: int) -> list[str]:
    if extension == ".xlsx":
        return parse_xlsx_emails(file_bytes, max_rows=max_rows)
    return parse_text_emails(file_bytes, extension=extension, max_rows=max_rows)


def parse_text_emails(file_bytes: bytes, extension: str, max_rows: int) -> list[str]:
    text_value = decode_file_bytes(file_bytes)
    lines = [line for line in text_value.splitlines() if line.strip()]
    if not lines:
        return []

    if extension == ".txt" and not contains_delimiter(lines[:25]):
        rows = [[line.strip()] for line in lines[:max_rows]]
    else:
        delimiter = detect_delimiter("\n".join(lines[:25]))
        reader = csv.reader(io.StringIO(text_value), delimiter=delimiter)
        rows = list(limit_rows(reader, max_rows=max_rows))

    emails = extract_emails_from_rows(rows)
    return emails


def parse_xlsx_emails(file_bytes: bytes, max_rows: int) -> list[str]:
    if load_workbook is None:
        raise ValueError("XLSX support requires openpyxl")

    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            rows: list[list[str]] = []
            for index, row in enumerate(worksheet.iter_rows(values_only=True)):
                if index >= max_rows:
                    break
                rows.append([normalize_cell(value) for value in row])

            emails = extract_emails_from_rows(rows)
            if emails:
                return emails
    finally:
        workbook.close()

    return []


def extract_emails_from_rows(rows: list[list[str]]) -> list[str]:
    cleaned_rows = [trim_row(row) for row in rows if any(normalize_cell(value) for value in row)]
    if not cleaned_rows:
        return []

    column_index, has_header = detect_email_column(cleaned_rows)
    if column_index is None:
        return []

    start_index = 1 if has_header else 0
    emails: list[str] = []
    for row in cleaned_rows[start_index:]:
        if column_index >= len(row):
            continue
        candidate = row[column_index].strip()
        if not candidate:
            continue
        if "@" in candidate:
            emails.append(candidate)

    return emails


def detect_email_column(rows: list[list[str]]) -> tuple[int | None, bool]:
    header = rows[0]
    for index, value in enumerate(header):
        if "email" in value.lower():
            return index, True

    max_columns = max(len(row) for row in rows)
    best_column: int | None = None
    best_score = 0
    sample_rows = rows[:100]

    for column_index in range(max_columns):
        score = 0
        for row in sample_rows:
            if column_index >= len(row):
                continue
            candidate = row[column_index].strip().lower()
            if not candidate:
                continue
            if EMAIL_REGEX.fullmatch(candidate):
                score += 3
            elif "@" in candidate:
                score += 1

        if score > best_score:
            best_score = score
            best_column = column_index

    return best_column, False


def decode_file_bytes(file_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("unable to decode uploaded file")


def limit_rows(rows: Any, max_rows: int) -> list[list[str]]:
    limited_rows: list[list[str]] = []
    for index, row in enumerate(rows):
        if index >= max_rows:
            break
        limited_rows.append([normalize_cell(value) for value in row])
    return limited_rows


def contains_delimiter(lines: list[str]) -> bool:
    return any(any(delimiter in line for delimiter in [",", ";", "\t", "|"]) for line in lines)


def detect_delimiter(sample: str) -> str:
    counts = {delimiter: sample.count(delimiter) for delimiter in [",", ";", "\t", "|"]}
    delimiter, count = max(counts.items(), key=lambda item: item[1])
    return delimiter if count > 0 else ","


def trim_row(row: list[str]) -> list[str]:
    return [normalize_cell(value) for value in row]


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def resolve_user_id(app: Flask, services: BackendServices) -> uuid.UUID:
    explicit_user_id = (
        request.headers.get("X-User-Id")
        or request.form.get("user_id")
    )
    if explicit_user_id:
        try:
            return uuid.UUID(str(explicit_user_id))
        except ValueError as exc:
            raise ValueError("invalid user_id") from exc

    api_key = request.headers.get("X-API-Key")
    if api_key:
        with services.session_factory() as session:
            user_id = session.execute(select(User.id).where(User.api_key == api_key)).scalar_one_or_none()
        if user_id is None:
            raise ValueError("invalid api key")
        return user_id

    default_user_id = app.config.get("DEFAULT_USER_ID")
    if default_user_id:
        try:
            return uuid.UUID(str(default_user_id))
        except ValueError as exc:
            raise ValueError("invalid DEFAULT_USER_ID") from exc

    raise ValueError("user_id or X-API-Key is required")


def parse_job_id(raw_job_id: str | None) -> uuid.UUID | None:
    if not raw_job_id:
        return None
    try:
        return uuid.UUID(raw_job_id)
    except ValueError:
        return None


def build_progress_payload(
    total: int,
    processed: int,
    valid: int,
    risky: int,
    invalid: int,
    status: str,
) -> dict[str, Any]:
    percent = int((processed / total) * 100) if total else 100
    return {
        "percent": percent,
        "processed": processed,
        "total": total,
        "valid": valid,
        "risky": risky,
        "invalid": invalid,
        "status": status,
    }


def progress_response_payload(payload: dict[str, Any]) -> dict[str, Any]:
    return {
        "percent": int(payload["percent"]),
        "processed": int(payload["processed"]),
        "total": int(payload["total"]),
        "valid": int(payload["valid"]),
        "risky": int(payload["risky"]),
        "invalid": int(payload["invalid"]),
        "status": payload.get("status", "running"),
    }


def update_progress_counts(progress: dict[str, Any], status: str) -> None:
    progress["processed"] += 1
    if status == "valid":
        progress["valid"] += 1
    elif status == "risky":
        progress["risky"] += 1
    else:
        progress["invalid"] += 1

    total = progress["total"]
    progress["percent"] = int((progress["processed"] / total) * 100) if total else 100


def sync_job_progress(
    session: Session,
    job_id: uuid.UUID,
    progress: dict[str, Any],
    status: str | None = None,
    r2_file_key: str | None = None,
    finished_at: datetime | None = None,
) -> None:
    values: dict[str, Any] = {
        "processed": progress["processed"],
        "valid_count": progress["valid"],
        "risky_count": progress["risky"],
        "invalid_count": progress["invalid"],
    }
    if status is not None:
        values["status"] = status
    if r2_file_key is not None:
        values["r2_file_key"] = r2_file_key
    if finished_at is not None:
        values["finished_at"] = finished_at

    session.execute(update(Job).where(Job.id == job_id).values(**values))
    session.commit()


def set_job_status(session: Session, job_id: uuid.UUID, status: str) -> None:
    session.execute(update(Job).where(Job.id == job_id).values(status=status))
    session.commit()


def store_progress(redis_client: redis.Redis, job_id: uuid.UUID, payload: dict[str, Any]) -> None:
    redis_client.setex(progress_key(job_id), PROGRESS_TTL_SECONDS, json.dumps(payload))


def load_progress(redis_client: redis.Redis, job_id: uuid.UUID) -> dict[str, Any] | None:
    raw_value = redis_client.get(progress_key(job_id))
    if not raw_value:
        return None

    if isinstance(raw_value, bytes):
        raw_value = raw_value.decode("utf-8")
    return json.loads(raw_value)


def load_cached_email_result(redis_client: redis.Redis, email: str) -> ValidationResult | None:
    raw_value = redis_client.get(email_cache_key(email))
    if not raw_value:
        return None

    if isinstance(raw_value, bytes):
        raw_value = raw_value.decode("utf-8")
    payload = json.loads(raw_value)
    result = ValidationResult(
        email=payload["email"],
        status=payload["status"],
        reason=payload["reason"],
    )
    if should_bypass_cached_result(result):
        return None
    return result


def store_cached_email_result(redis_client: redis.Redis, result: ValidationResult) -> None:
    payload = {
        "email": result.email,
        "status": result.status,
        "reason": result.reason,
    }
    ttl_seconds = (
        TRANSIENT_EMAIL_RESULT_TTL_SECONDS
        if should_bypass_cached_result(result)
        else EMAIL_RESULT_TTL_SECONDS
    )
    redis_client.setex(
        email_cache_key(result.email),
        ttl_seconds,
        json.dumps(payload),
    )


def upsert_email_cache(session: Session, result: ValidationResult) -> None:
    statement = pg_insert(EmailCache.__table__).values(
        email=result.email,
        status=result.status,
        reason=result.reason,
        checked_at=datetime.now(timezone.utc),
    )
    statement = statement.on_conflict_do_update(
        index_elements=[EmailCache.__table__.c.email],
        set_={
            "status": result.status,
            "reason": result.reason,
            "checked_at": datetime.now(timezone.utc),
        },
    )
    session.execute(statement)


def render_results_csv(results: list[dict[str, str]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=["email", "status", "reason"])
    writer.writeheader()
    writer.writerows(results)
    return buffer.getvalue().encode("utf-8")


def upload_r2_object(services: BackendServices, key: str, body: bytes) -> None:
    services.r2_client.put_object(
        Bucket=services.r2_bucket,
        Key=key,
        Body=body,
        ContentType="text/csv",
    )


def fetch_r2_object(services: BackendServices, key: str) -> bytes:
    response = services.r2_client.get_object(Bucket=services.r2_bucket, Key=key)
    return response["Body"].read()


def filter_result_csv(source_bytes: bytes, download_type: str) -> bytes:
    source_text = source_bytes.decode("utf-8")
    reader = csv.DictReader(io.StringIO(source_text))

    filtered_rows: list[dict[str, str]] = []
    for row in reader:
        if should_include_row(row, download_type):
            filtered_rows.append(
                {
                    "email": row.get("email", ""),
                    "status": row.get("status", ""),
                    "reason": row.get("reason", ""),
                }
            )

    return render_results_csv(filtered_rows)


def should_include_row(row: dict[str, str], download_type: str) -> bool:
    status = (row.get("status") or "").strip().lower()
    if download_type == "full":
        return True
    if download_type == "safe":
        return status == "valid"
    if download_type == "risky":
        return status == "risky"
    if download_type == "invalid":
        return status == "invalid"
    return False


def is_cancel_requested(redis_client: redis.Redis, job_id: uuid.UUID) -> bool:
    return bool(redis_client.get(cancel_key(job_id)))


def fetch_probe_health(url: str) -> dict[str, Any]:
    try:
        with urllib_request.urlopen(url, timeout=5) as response:
            payload = response.read().decode("utf-8")
        return json.loads(payload)
    except (urllib_error.URLError, TimeoutError, json.JSONDecodeError) as exc:
        return {"status": "error", "reason": str(exc)}


def should_bypass_cached_result(result: ValidationResult) -> bool:
    return result.status == "risky" and result.reason in TRANSIENT_RISKY_REASONS


def email_cache_key(email: str) -> str:
    return f"{EMAIL_CACHE_PREFIX}{email.strip().lower()}"


def progress_key(job_id: uuid.UUID) -> str:
    return f"{JOB_PROGRESS_PREFIX}{job_id}"


def cancel_key(job_id: uuid.UUID) -> str:
    return f"{JOB_CANCEL_PREFIX}{job_id}"


def get_services(app: Flask) -> BackendServices:
    return app.extensions["backend_services"]


def require_env(name: str) -> str:
    value = os.getenv(name)
    if value:
        return value
    raise RuntimeError(f"{name} is required")


app = create_app()


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "5000")))
